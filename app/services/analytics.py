from datetime import date, datetime, timedelta
from io import BytesIO
from typing import List, Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.infra.postgres.models.device import Device
from app.infra.postgres.models.payment import Payment
from app.infra.postgres.models.role import Role
from app.infra.postgres.models.user import User
from app.schemas.analytics import AnalyticsResponse, DailyAnalytics


class AnalyticsService:
    @staticmethod
    async def get_analytics_by_date_range(
        start_date: date, end_date: date = None, store_id: Optional[UUID] = None
    ) -> AnalyticsResponse:
        """
        Get analytics data for a date range with daily breakdowns.
        If end_date is None, use current date.
        Returns daily counts and totals for the date range.
        """
        if end_date is None:
            end_date = date.today()

        # Ensure start_date is not after end_date
        if start_date > end_date:
            start_date, end_date = end_date, start_date

        # Convert dates to datetime for comparison
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())

        # Initialize daily data list
        daily_data = []
        current_date = start_date
        
        # Calculate totals and daily breakdowns
        total_customers = 0
        total_devices = 0
        total_payments = 0.0
        total_vendors = 0
        
        while current_date <= end_date:
            current_datetime_start = datetime.combine(current_date, datetime.min.time())
            current_datetime_end = datetime.combine(current_date, datetime.max.time())
            
            # Get customers count for the day
            customer_query = User.filter(
                role__name="Cliente",
                created_at__gte=current_datetime_start,
                created_at__lte=current_datetime_end
            )
            
            # Filter by store_id if provided
            if store_id:
                customer_query = customer_query.filter(store_id=store_id)
                
            customers = await customer_query.count()
            
            # Get vendors count for the day
            vendor_query = User.filter(
                role__name="Vendedor",
                created_at__gte=current_datetime_start,
                created_at__lte=current_datetime_end
            )
            
            # Filter by store_id if provided
            if store_id:
                vendor_query = vendor_query.filter(store_id=store_id)
                
            vendors = await vendor_query.count()
            
            # Get devices count for the day
            device_query = Device.filter(
                created_at__gte=current_datetime_start,
                created_at__lte=current_datetime_end
            )
            
            # Filter by store_id if provided - devices are linked to users through enrolment
            if store_id:
                device_query = device_query.filter(enrolment__user__store_id=store_id)
                
            devices = await device_query.count()
            
            # Get payments value for the day
            payment_query = Payment.filter(
                date__gte=current_datetime_start,
                date__lte=current_datetime_end
            )
            
            # Filter by store_id if provided
            if store_id:
                payment_query = payment_query.filter(plan__user__store_id=store_id)
                
            payments = await payment_query.all()
            payments_value = sum(float(payment.value) for payment in payments)
            
            # Add to totals
            total_customers += customers
            total_vendors += vendors
            total_devices += devices
            total_payments += payments_value
            
            # Add daily data
            daily_data.append(DailyAnalytics(
                date=current_date,
                customers=customers,
                devices=devices,
                payments=payments_value,
                vendors=vendors
            ))
            
            current_date += timedelta(days=1)
        
        return AnalyticsResponse(
            total_customers=total_customers,
            total_devices=total_devices,
            total_payments=total_payments,
            total_vendors=total_vendors,
            daily_data=daily_data
        )

    @staticmethod
    async def generate_analytics_excel(
        start_date: date, end_date: date = None, store_id: Optional[UUID] = None
    ) -> BytesIO:
        """
        Generate Excel file with detailed analytics data for a date range.
        """
        if end_date is None:
            end_date = date.today()

        # Ensure start_date is not after end_date
        if start_date > end_date:
            start_date, end_date = end_date, start_date

        # Convert dates to datetime for comparison
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Report"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center")

        # Get detailed data
        customer_query = User.filter(
            created_at__gte=start_datetime,
            created_at__lte=end_datetime,
            role__name="Cliente"
        )
        
        # Filter by store_id if provided
        if store_id:
            customer_query = customer_query.filter(store_id=store_id)
            
        customers = await customer_query.prefetch_related("role", "city")

        vendor_query = User.filter(
            created_at__gte=start_datetime,
            created_at__lte=end_datetime,
            role__name="Vendedor"
        )
        
        # Filter by store_id if provided
        if store_id:
            vendor_query = vendor_query.filter(store_id=store_id)
            
        vendors = await vendor_query.prefetch_related("role", "city")

        device_query = Device.filter(
            created_at__gte=start_datetime,
            created_at__lte=end_datetime
        )
        
        # Filter by store_id if provided
        if store_id:
            device_query = device_query.filter(enrolment__user__store_id=store_id)
            
        devices = await device_query.prefetch_related("enrolment")

        payment_query = Payment.filter(
            date__gte=start_datetime,
            date__lte=end_datetime
        )
        
        # Filter by store_id if provided
        if store_id:
            payment_query = payment_query.filter(plan__user__store_id=store_id)
            
        payments = await payment_query.prefetch_related("device", "plan")

        # Summary section
        ws["A1"] = "REPORTE DE ANALYTICS"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A2"] = f"Período: {start_date.strftime('%Y-%m-%d')} al {end_date.strftime('%Y-%m-%d')}"
        ws["A2"].font = Font(bold=True)

        # Summary totals
        ws["A4"] = "RESUMEN"
        ws["A4"].font = header_font
        ws["A4"].fill = header_fill
        ws["B4"].fill = header_fill
        
        ws["A5"] = "Total Clientes:"
        ws["B5"] = len(customers)
        ws["A6"] = "Total Vendedores:"
        ws["B6"] = len(vendors)
        ws["A7"] = "Total Dispositivos:"
        ws["B7"] = len(devices)
        ws["A8"] = "Total Pagos:"
        ws["B8"] = sum(float(payment.value) for payment in payments)

        # Customers detail
        current_row = 11
        ws[f"A{current_row}"] = "DETALLE DE CLIENTES"
        ws[f"A{current_row}"].font = header_font
        ws[f"A{current_row}"].fill = header_fill
        for col in range(1, 7):
            ws[f"{get_column_letter(col)}{current_row}"].fill = header_fill
        
        current_row += 1
        headers = ["DNI", "Nombre Completo", "Email", "Teléfono", "Ciudad", "Fecha Creación"]
        for col, header in enumerate(headers, 1):
            cell = ws[f"{get_column_letter(col)}{current_row}"]
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        current_row += 1
        for customer in customers:
            ws[f"A{current_row}"] = customer.dni
            ws[f"B{current_row}"] = f"{customer.first_name} {customer.last_name}"
            ws[f"C{current_row}"] = customer.email
            ws[f"D{current_row}"] = f"{customer.prefix}{customer.phone}"
            ws[f"E{current_row}"] = customer.city.name if customer.city else "N/A"
            ws[f"F{current_row}"] = customer.created_at.strftime('%Y-%m-%d %H:%M')
            current_row += 1

        # Vendors detail
        current_row += 2
        ws[f"A{current_row}"] = "DETALLE DE VENDEDORES"
        ws[f"A{current_row}"].font = header_font
        ws[f"A{current_row}"].fill = header_fill
        for col in range(1, 7):
            ws[f"{get_column_letter(col)}{current_row}"].fill = header_fill
        
        current_row += 1
        for col, header in enumerate(headers, 1):
            cell = ws[f"{get_column_letter(col)}{current_row}"]
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        current_row += 1
        for vendor in vendors:
            ws[f"A{current_row}"] = vendor.dni
            ws[f"B{current_row}"] = f"{vendor.first_name} {vendor.last_name}"
            ws[f"C{current_row}"] = vendor.email
            ws[f"D{current_row}"] = f"{vendor.prefix}{vendor.phone}"
            ws[f"E{current_row}"] = vendor.city.name if vendor.city else "N/A"
            ws[f"F{current_row}"] = vendor.created_at.strftime('%Y-%m-%d %H:%M')
            current_row += 1

        # Devices detail
        current_row += 2
        ws[f"A{current_row}"] = "DETALLE DE DISPOSITIVOS"
        ws[f"A{current_row}"].font = header_font
        ws[f"A{current_row}"].fill = header_fill
        for col in range(1, 7):
            ws[f"{get_column_letter(col)}{current_row}"].fill = header_fill
        
        current_row += 1
        device_headers = ["IMEI", "Nombre", "Marca", "Modelo", "Estado", "Fecha Creación"]
        for col, header in enumerate(device_headers, 1):
            cell = ws[f"{get_column_letter(col)}{current_row}"]
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        current_row += 1
        for device in devices:
            ws[f"A{current_row}"] = device.imei
            ws[f"B{current_row}"] = device.name
            ws[f"C{current_row}"] = device.brand
            ws[f"D{current_row}"] = device.model
            ws[f"E{current_row}"] = device.state.value
            ws[f"F{current_row}"] = device.created_at.strftime('%Y-%m-%d %H:%M')
            current_row += 1

        # Payments detail
        current_row += 2
        ws[f"A{current_row}"] = "DETALLE DE PAGOS"
        ws[f"A{current_row}"].font = header_font
        ws[f"A{current_row}"].fill = header_fill
        for col in range(1, 6):
            ws[f"{get_column_letter(col)}{current_row}"].fill = header_fill
        
        current_row += 1
        payment_headers = ["Referencia", "Valor", "Método", "Estado", "Fecha"]
        for col, header in enumerate(payment_headers, 1):
            cell = ws[f"{get_column_letter(col)}{current_row}"]
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        current_row += 1
        for payment in payments:
            ws[f"A{current_row}"] = payment.reference
            ws[f"B{current_row}"] = float(payment.value)
            ws[f"C{current_row}"] = payment.method
            ws[f"D{current_row}"] = payment.state.value
            ws[f"E{current_row}"] = payment.date.strftime('%Y-%m-%d %H:%M')
            current_row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer


analytics_service = AnalyticsService()
